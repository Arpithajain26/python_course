import openpyxl as xl
from openpyxl.chart import BarChart, Reference

# Load workbook
wb = xl.load_workbook("Book1.xlsx")
sheet = wb["Sheet1"]

# Calculate corrected prices
for row in range(2, sheet.max_row + 1):
    cell = sheet.cell(row, 3)  # Column C (Price)

    # Remove '$' and convert to float
    price = float(str(cell.value).replace("$", ""))

    corrected_price = price * 0.9

    # Write corrected price to Column D
    sheet.cell(row, 4).value = corrected_price

# Create reference for the chart (Column D)
values = Reference(
    sheet,
    min_col=4,
    min_row=2,
    max_row=sheet.max_row
)

# Create bar chart
chart = BarChart()
chart.add_data(values)

# Add chart to worksheet
sheet.add_chart(chart, "F2")

# Save workbook
wb.save("transactions2.xlsx")

print("Workbook updated successfully!")